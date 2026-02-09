"""Use Case pour l'export comptable.

FIN-13: Export comptable - Generation CSV et Excel pour la comptabilite.
"""

import csv
import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import List

from ...domain.repositories.budget_repository import BudgetRepository
from ...domain.repositories.achat_repository import AchatRepository
from ...domain.repositories.situation_repository import SituationRepository
from ...domain.repositories.facture_repository import FactureRepository
from ...domain.repositories.fournisseur_repository import FournisseurRepository
from ...domain.repositories.lot_budgetaire_repository import LotBudgetaireRepository
from ..dtos.export_dtos import ExportComptableDTO, LigneExportComptableDTO
from shared.domain.calcul_financier import calculer_tva, arrondir_montant

logger = logging.getLogger(__name__)


class ExportComptableError(Exception):
    """Exception pour les erreurs d'export comptable."""

    def __init__(self, message: str = "Erreur lors de l'export comptable"):
        self.message = message
        super().__init__(self.message)


class ExportComptableUseCase:
    """Cas d'utilisation : Generer un export comptable pour un chantier.

    Collecte les achats factures, les situations validees/facturees et
    les factures client pour generer un export CSV ou Excel.

    Attributes:
        budget_repo: Repository pour les budgets.
        achat_repo: Repository pour les achats.
        situation_repo: Repository pour les situations.
        facture_repo: Repository pour les factures.
        fournisseur_repo: Repository pour les fournisseurs.
        lot_repo: Repository pour les lots budgetaires.
    """

    def __init__(
        self,
        budget_repo: BudgetRepository,
        achat_repo: AchatRepository,
        situation_repo: SituationRepository,
        facture_repo: FactureRepository,
        fournisseur_repo: FournisseurRepository,
        lot_repo: LotBudgetaireRepository,
    ):
        """Initialise le use case.

        Args:
            budget_repo: Repository budget (interface).
            achat_repo: Repository achat (interface).
            situation_repo: Repository situation (interface).
            facture_repo: Repository facture (interface).
            fournisseur_repo: Repository fournisseur (interface).
            lot_repo: Repository lot budgetaire (interface).
        """
        self.budget_repo = budget_repo
        self.achat_repo = achat_repo
        self.situation_repo = situation_repo
        self.facture_repo = facture_repo
        self.fournisseur_repo = fournisseur_repo
        self.lot_repo = lot_repo

    def _generer_code_analytique(
        self, chantier_id: int, code_lot: str | None = None
    ) -> str:
        """Genere un code analytique pour une ligne d'export.

        Args:
            chantier_id: L'ID du chantier.
            code_lot: Le code du lot budgetaire (optionnel).

        Returns:
            Le code analytique formate.
        """
        if code_lot:
            return f"CHANT-{chantier_id:03d}-LOT-{code_lot}"
        return f"CHANT-{chantier_id:03d}"

    def _get_fournisseur_nom(self, fournisseur_id: int | None) -> str:
        """Recupere le nom d'un fournisseur.

        Args:
            fournisseur_id: L'ID du fournisseur.

        Returns:
            Le nom du fournisseur ou une chaine vide.
        """
        if not fournisseur_id:
            return ""
        fournisseur = self.fournisseur_repo.find_by_id(fournisseur_id)
        return fournisseur.raison_sociale if fournisseur else ""

    def _get_lot_code(self, lot_id: int | None) -> str | None:
        """Recupere le code d'un lot budgetaire.

        Args:
            lot_id: L'ID du lot budgetaire.

        Returns:
            Le code du lot ou None.
        """
        if not lot_id:
            return None
        lot = self.lot_repo.find_by_id(lot_id)
        return lot.code_lot if lot else None

    def execute(self, chantier_id: int) -> ExportComptableDTO:
        """Execute l'export comptable pour un chantier.

        Args:
            chantier_id: L'ID du chantier.

        Returns:
            ExportComptableDTO contenant toutes les lignes de l'export.

        Raises:
            ExportComptableError: Si le budget n'existe pas.
        """
        # Verifier que le budget existe
        budget = self.budget_repo.find_by_chantier_id(chantier_id)
        if not budget:
            raise ExportComptableError(
                f"Aucun budget pour le chantier {chantier_id}"
            )

        lignes: List[LigneExportComptableDTO] = []
        total_ht = Decimal("0")
        total_tva = Decimal("0")
        total_ttc = Decimal("0")

        # Cache fournisseurs
        fournisseurs_cache: dict[int, str] = {}
        lots_cache: dict[int, str | None] = {}

        # 1. Collecter les achats (statut facture uniquement)
        from ...domain.value_objects import StatutAchat
        achats = self.achat_repo.find_by_chantier(
            chantier_id=chantier_id,
            statut=StatutAchat.FACTURE,
            limit=10000,
            offset=0,
        )
        for achat in achats:
            montant_ht = arrondir_montant(achat.quantite * achat.prix_unitaire_ht)
            montant_tva = calculer_tva(montant_ht, achat.taux_tva)
            montant_ttc = montant_ht + montant_tva

            # Cache fournisseur
            if achat.fournisseur_id and achat.fournisseur_id not in fournisseurs_cache:
                fournisseurs_cache[achat.fournisseur_id] = self._get_fournisseur_nom(
                    achat.fournisseur_id
                )

            # Cache lot
            if achat.lot_budgetaire_id and achat.lot_budgetaire_id not in lots_cache:
                lots_cache[achat.lot_budgetaire_id] = self._get_lot_code(
                    achat.lot_budgetaire_id
                )

            code_lot = lots_cache.get(achat.lot_budgetaire_id) if achat.lot_budgetaire_id else None
            tiers = fournisseurs_cache.get(achat.fournisseur_id, "") if achat.fournisseur_id else ""

            lignes.append(
                LigneExportComptableDTO(
                    date=(
                        achat.date_commande.isoformat()
                        if achat.date_commande
                        else ""
                    ),
                    type_document="achat",
                    numero=achat.numero_facture or f"ACH-{achat.id}",
                    tiers=tiers,
                    montant_ht=str(montant_ht),
                    montant_tva=str(montant_tva),
                    montant_ttc=str(montant_ttc),
                    code_analytique=self._generer_code_analytique(
                        chantier_id, code_lot
                    ),
                    libelle=achat.libelle,
                    reference_chantier=f"CHANT-{chantier_id:03d}",
                    statut=achat.statut.value if hasattr(achat.statut, 'value') else str(achat.statut),
                )
            )
            total_ht += montant_ht
            total_tva += montant_tva
            total_ttc += montant_ttc

        # 2. Collecter les situations (validee/facturee)
        situations = self.situation_repo.find_by_chantier_id(chantier_id)
        for situation in situations:
            if situation.statut not in ("validee", "facturee"):
                continue

            montant_ht = situation.montant_periode_ht
            montant_tva = calculer_tva(montant_ht, situation.taux_tva)
            montant_ttc = montant_ht + montant_tva

            lignes.append(
                LigneExportComptableDTO(
                    date=(
                        situation.periode_fin.isoformat()
                        if situation.periode_fin
                        else ""
                    ),
                    type_document="situation",
                    numero=situation.numero,
                    tiers="Greg Construction",
                    montant_ht=str(montant_ht),
                    montant_tva=str(montant_tva),
                    montant_ttc=str(montant_ttc),
                    code_analytique=self._generer_code_analytique(chantier_id),
                    libelle=f"Situation {situation.numero}",
                    reference_chantier=f"CHANT-{chantier_id:03d}",
                    statut=situation.statut,
                )
            )
            total_ht += montant_ht
            total_tva += montant_tva
            total_ttc += montant_ttc

        # 3. Collecter les factures
        factures = self.facture_repo.find_by_chantier_id(chantier_id)
        for facture in factures:
            if facture.statut in ("brouillon", "annulee"):
                continue

            lignes.append(
                LigneExportComptableDTO(
                    date=(
                        facture.date_emission.isoformat()
                        if facture.date_emission
                        else ""
                    ),
                    type_document="facture",
                    numero=facture.numero_facture,
                    tiers="Greg Construction",
                    montant_ht=str(facture.montant_ht),
                    montant_tva=str(facture.montant_tva),
                    montant_ttc=str(facture.montant_ttc),
                    code_analytique=self._generer_code_analytique(chantier_id),
                    libelle=f"Facture {facture.numero_facture} ({facture.type_facture})",
                    reference_chantier=f"CHANT-{chantier_id:03d}",
                    statut=facture.statut,
                )
            )
            total_ht += facture.montant_ht
            total_tva += facture.montant_tva
            total_ttc += facture.montant_ttc

        # Trier par date
        lignes.sort(key=lambda l: l.date)

        dto = ExportComptableDTO(
            chantier_id=chantier_id,
            nom_chantier=f"Chantier {chantier_id}",
            date_export=datetime.utcnow().isoformat(),
            lignes=lignes,
            totaux={
                "total_ht": str(arrondir_montant(total_ht)),
                "total_tva": str(arrondir_montant(total_tva)),
                "total_ttc": str(arrondir_montant(total_ttc)),
            },
        )

        logger.info(
            "Export comptable genere pour chantier %s: %d lignes",
            chantier_id, len(lignes),
        )
        return dto

    def to_csv(self, dto: ExportComptableDTO) -> str:
        """Genere un fichier CSV a partir du DTO d'export.

        Args:
            dto: Le DTO d'export comptable.

        Returns:
            Le contenu CSV en UTF-8-BOM avec separateur point-virgule.
        """
        output = io.StringIO()
        # BOM UTF-8 pour Excel
        output.write("\ufeff")

        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

        # En-tetes
        writer.writerow([
            "Date",
            "Type Document",
            "Numero",
            "Tiers",
            "Montant HT",
            "Montant TVA",
            "Montant TTC",
            "Code Analytique",
            "Libelle",
            "Reference Chantier",
            "Statut",
        ])

        # Lignes
        for ligne in dto.lignes:
            writer.writerow([
                ligne.date,
                ligne.type_document,
                ligne.numero,
                ligne.tiers,
                ligne.montant_ht,
                ligne.montant_tva,
                ligne.montant_ttc,
                ligne.code_analytique,
                ligne.libelle,
                ligne.reference_chantier,
                ligne.statut,
            ])

        # Ligne totaux
        writer.writerow([])
        writer.writerow([
            "", "", "", "TOTAUX",
            dto.totaux.get("total_ht", "0"),
            dto.totaux.get("total_tva", "0"),
            dto.totaux.get("total_ttc", "0"),
            "", "", "", "",
        ])

        return output.getvalue()

    def to_xlsx(self, dto: ExportComptableDTO) -> bytes:
        """Genere un fichier Excel a partir du DTO d'export.

        Args:
            dto: Le DTO d'export comptable.

        Returns:
            Le contenu Excel en bytes.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Export Comptable"

        # En-tetes en gras
        headers = [
            "Date",
            "Type Document",
            "Numero",
            "Tiers",
            "Montant HT",
            "Montant TVA",
            "Montant TTC",
            "Code Analytique",
            "Libelle",
            "Reference Chantier",
            "Statut",
        ]
        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

        # Donnees
        for row_idx, ligne in enumerate(dto.lignes, 2):
            ws.cell(row=row_idx, column=1, value=ligne.date)
            ws.cell(row=row_idx, column=2, value=ligne.type_document)
            ws.cell(row=row_idx, column=3, value=ligne.numero)
            ws.cell(row=row_idx, column=4, value=ligne.tiers)
            ws.cell(row=row_idx, column=5, value=ligne.montant_ht)
            ws.cell(row=row_idx, column=6, value=ligne.montant_tva)
            ws.cell(row=row_idx, column=7, value=ligne.montant_ttc)
            ws.cell(row=row_idx, column=8, value=ligne.code_analytique)
            ws.cell(row=row_idx, column=9, value=ligne.libelle)
            ws.cell(row=row_idx, column=10, value=ligne.reference_chantier)
            ws.cell(row=row_idx, column=11, value=ligne.statut)

        # Ligne totaux
        totals_row = len(dto.lignes) + 3
        ws.cell(row=totals_row, column=4, value="TOTAUX").font = bold_font
        ws.cell(
            row=totals_row, column=5,
            value=dto.totaux.get("total_ht", "0"),
        ).font = bold_font
        ws.cell(
            row=totals_row, column=6,
            value=dto.totaux.get("total_tva", "0"),
        ).font = bold_font
        ws.cell(
            row=totals_row, column=7,
            value=dto.totaux.get("total_ttc", "0"),
        ).font = bold_font

        # Auto-width des colonnes
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Metadata
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.cell(row=1, column=1, value="Chantier ID").font = bold_font
        ws_meta.cell(row=1, column=2, value=dto.chantier_id)
        ws_meta.cell(row=2, column=1, value="Nom Chantier").font = bold_font
        ws_meta.cell(row=2, column=2, value=dto.nom_chantier)
        ws_meta.cell(row=3, column=1, value="Date Export").font = bold_font
        ws_meta.cell(row=3, column=2, value=dto.date_export)
        ws_meta.cell(row=4, column=1, value="Nombre Lignes").font = bold_font
        ws_meta.cell(row=4, column=2, value=len(dto.lignes))

        # Sauvegarder en bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            "Export Excel genere pour chantier %s: %d lignes",
            dto.chantier_id, len(dto.lignes),
        )
        return output.read()
