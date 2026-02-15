"""Use Case pour l'import DPGF automatique.

DEV-21: Import fichier DPGF Excel/CSV.
Parse les colonnes: lot, description, unite, quantite, PU.
Cree automatiquement les lots et lignes dans le devis.
"""

import csv
import io
import logging
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO, Dict, List, Optional, Tuple

from ...domain.entities.devis import Devis
from ...domain.entities.journal_devis import JournalDevis
from ...domain.entities.lot_devis import LotDevis
from ...domain.entities.ligne_devis import LigneDevis
from ...domain.value_objects import StatutDevis, UniteArticle
from ...domain.repositories.devis_repository import DevisRepository
from ...domain.repositories.lot_devis_repository import LotDevisRepository
from ...domain.repositories.ligne_devis_repository import LigneDevisRepository
from ...domain.repositories.journal_devis_repository import JournalDevisRepository
from ..dtos.dpgf_dtos import (
    DPGFColumnMappingDTO,
    ImportDPGFResultDTO,
    LigneDPGFDTO,
)

logger = logging.getLogger(__name__)


class DPGFImportError(Exception):
    """Erreur lors de l'import DPGF."""

    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


class DPGFFormatError(DPGFImportError):
    """Format de fichier DPGF non reconnu ou invalide."""
    pass


class DevisNonImportableError(DPGFImportError):
    """Le devis n'est pas dans un etat permettant l'import DPGF."""
    pass


# Mapping des unites DPGF courantes vers UniteArticle
_UNITE_MAPPING: Dict[str, UniteArticle] = {
    "u": UniteArticle.U,
    "unite": UniteArticle.U,
    "ens": UniteArticle.ENS,
    "ensemble": UniteArticle.ENS,
    "m": UniteArticle.ML,
    "ml": UniteArticle.ML,
    "m2": UniteArticle.M2,
    "m\u00b2": UniteArticle.M2,
    "m3": UniteArticle.M3,
    "m\u00b3": UniteArticle.M3,
    "kg": UniteArticle.KG,
    "t": UniteArticle.T,
    "tonne": UniteArticle.T,
    "h": UniteArticle.HEURE,
    "heure": UniteArticle.HEURE,
    "heures": UniteArticle.HEURE,
    "j": UniteArticle.JOUR,
    "jour": UniteArticle.JOUR,
    "jours": UniteArticle.JOUR,
    "ft": UniteArticle.FORFAIT,
    "fft": UniteArticle.FORFAIT,
    "forfait": UniteArticle.FORFAIT,
    "l": UniteArticle.L,
    "litre": UniteArticle.L,
}


def _parse_unite(raw: str) -> UniteArticle:
    """Parse une unite depuis un fichier DPGF.

    Args:
        raw: La valeur brute de la colonne unite.

    Returns:
        L'UniteArticle correspondante, ou U par defaut.
    """
    cleaned = raw.strip().lower()
    return _UNITE_MAPPING.get(cleaned, UniteArticle.U)


def _parse_decimal(raw: str, champ: str, ligne_num: int) -> Tuple[Decimal, Optional[str]]:
    """Parse une valeur decimale depuis un fichier DPGF.

    Args:
        raw: La valeur brute.
        champ: Nom du champ (pour message d'erreur).
        ligne_num: Numero de ligne (pour message d'erreur).

    Returns:
        Tuple (valeur, erreur_ou_none).
    """
    cleaned = raw.strip().replace(",", ".").replace(" ", "").replace("\u00a0", "")
    if not cleaned or cleaned == "-":
        return Decimal("0"), None
    try:
        val = Decimal(cleaned)
        if val < 0:
            return Decimal("0"), f"Ligne {ligne_num}: {champ} negatif ({cleaned}), mis a 0"
        return val, None
    except InvalidOperation:
        return Decimal("0"), f"Ligne {ligne_num}: {champ} invalide ({raw!r}), mis a 0"


class ImportDPGFUseCase:
    """Use Case pour importer un fichier DPGF dans un devis.

    DEV-21: Parse un fichier Excel (.xlsx) ou CSV et cree
    automatiquement les lots et lignes dans le devis cible.

    Le devis doit etre en statut brouillon ou en negociation.
    L'import ne supprime PAS les lots/lignes existants.
    """

    def __init__(
        self,
        devis_repository: DevisRepository,
        lot_repository: LotDevisRepository,
        ligne_repository: LigneDevisRepository,
        journal_repository: JournalDevisRepository,
    ):
        self._devis_repository = devis_repository
        self._lot_repository = lot_repository
        self._ligne_repository = ligne_repository
        self._journal_repository = journal_repository

    def execute(
        self,
        devis_id: int,
        file_content: bytes,
        filename: str,
        imported_by: int,
        mapping: Optional[DPGFColumnMappingDTO] = None,
    ) -> ImportDPGFResultDTO:
        """Importe un fichier DPGF dans un devis.

        Args:
            devis_id: ID du devis cible.
            file_content: Contenu binaire du fichier.
            filename: Nom du fichier (pour detecter le format).
            imported_by: ID de l'utilisateur importateur.
            mapping: Mapping optionnel des colonnes.

        Returns:
            Resultat de l'import (lots/lignes crees, erreurs).

        Raises:
            DevisNotFoundError: Si le devis n'existe pas.
            DevisNonImportableError: Si le devis n'est pas modifiable.
            DPGFFormatError: Si le format du fichier est invalide.
        """
        from .devis_use_cases import DevisNotFoundError

        if mapping is None:
            mapping = DPGFColumnMappingDTO()

        # 1. Verification du devis
        devis = self._devis_repository.find_by_id(devis_id)
        if not devis:
            raise DevisNotFoundError(devis_id)

        if not devis.est_modifiable:
            raise DevisNonImportableError(
                f"Le devis {devis.numero} est en statut '{devis.statut.label}' "
                f"et ne peut pas recevoir un import DPGF"
            )

        # 2. Detection format et parsing
        filename_lower = filename.lower()
        if filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
            lignes_brutes = self._parse_excel(file_content, mapping)
        elif filename_lower.endswith(".csv"):
            lignes_brutes = self._parse_csv(file_content, mapping)
        else:
            raise DPGFFormatError(
                f"Format de fichier non supporte: {filename}. "
                f"Formats acceptes: .xlsx, .xls, .csv"
            )

        if not lignes_brutes:
            raise DPGFFormatError("Aucune ligne de donnees trouvee dans le fichier")

        # 3. Validation et parsing des lignes
        lignes_valides: List[LigneDPGFDTO] = []
        erreurs: List[dict] = []

        for idx, row in enumerate(lignes_brutes):
            ligne_num = idx + mapping.ligne_debut + 1  # Numero affiche (1-indexed)
            try:
                ligne_dto = self._parse_ligne(row, mapping, ligne_num)
                if ligne_dto:
                    lignes_valides.append(ligne_dto)
            except Exception as e:
                erreurs.append({
                    "ligne": ligne_num,
                    "message": str(e),
                })

        if not lignes_valides:
            raise DPGFFormatError(
                f"Aucune ligne valide trouvee. {len(erreurs)} erreurs de parsing."
            )

        # 4. Grouper par lot
        lignes_par_lot: Dict[str, List[LigneDPGFDTO]] = defaultdict(list)
        for ligne in lignes_valides:
            lignes_par_lot[ligne.lot].append(ligne)

        # 5. Determiner l'ordre de depart (apres les lots existants)
        lots_existants = self._lot_repository.find_by_devis(devis_id)
        ordre_lot = max((l.ordre for l in lots_existants), default=0) + 1

        # 6. Creer lots et lignes
        result = ImportDPGFResultDTO(devis_id=devis_id)
        lots_info = []

        for code_lot, lignes_lot in lignes_par_lot.items():
            # Creer le lot
            lot = LotDevis(
                devis_id=devis_id,
                code_lot=code_lot,
                libelle=f"Lot {code_lot}",
                ordre=ordre_lot,
                created_at=datetime.utcnow(),
                created_by=imported_by,
            )
            lot = self._lot_repository.save(lot)
            result.lots_crees += 1
            ordre_lot += 1

            # Creer les lignes du lot
            lignes_info = []
            for idx_ligne, ligne_dto in enumerate(lignes_lot):
                unite = _parse_unite(ligne_dto.unite)
                ligne = LigneDevis(
                    lot_devis_id=lot.id,
                    libelle=ligne_dto.description,
                    unite=unite,
                    quantite=ligne_dto.quantite,
                    prix_unitaire_ht=ligne_dto.prix_unitaire,
                    taux_tva=devis.taux_tva_defaut,
                    ordre=idx_ligne + 1,
                    total_ht=ligne_dto.quantite * ligne_dto.prix_unitaire,
                    created_at=datetime.utcnow(),
                    created_by=imported_by,
                )
                ligne = self._ligne_repository.save(ligne)
                result.lignes_creees += 1

                lignes_info.append({
                    "id": ligne.id,
                    "libelle": ligne.libelle,
                    "quantite": str(ligne.quantite),
                    "prix_unitaire_ht": str(ligne.prix_unitaire_ht),
                })

            lots_info.append({
                "id": lot.id,
                "code_lot": lot.code_lot,
                "libelle": lot.libelle,
                "nb_lignes": len(lignes_info),
                "lignes": lignes_info,
            })

        result.lots = lots_info
        result.lignes_ignorees = len(erreurs)
        result.erreurs = erreurs

        # 7. Journal
        self._journal_repository.save(
            JournalDevis(
                devis_id=devis_id,
                action="import_dpgf",
                details_json={
                    "message": (
                        f"Import DPGF depuis {filename}: "
                        f"{result.lots_crees} lots, {result.lignes_creees} lignes, "
                        f"{result.lignes_ignorees} ignorees"
                    ),
                    "type_modification": "import_dpgf",
                    "fichier": filename,
                    "lots_crees": result.lots_crees,
                    "lignes_creees": result.lignes_creees,
                    "lignes_ignorees": result.lignes_ignorees,
                },
                auteur_id=imported_by,
                created_at=datetime.utcnow(),
            )
        )

        return result

    def _parse_excel(
        self, content: bytes, mapping: DPGFColumnMappingDTO
    ) -> List[List[str]]:
        """Parse un fichier Excel (.xlsx/.xls).

        Returns:
            Liste de lignes, chaque ligne etant une liste de valeurs string.
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            if mapping.feuille < len(wb.sheetnames):
                ws = wb[wb.sheetnames[mapping.feuille]]
            else:
                ws = wb.active

            rows = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx < mapping.ligne_debut:
                    continue
                # Convertir en strings
                row_str = [str(cell) if cell is not None else "" for cell in row]
                # Ignorer les lignes completement vides
                if any(cell.strip() for cell in row_str):
                    rows.append(row_str)

            wb.close()
            return rows

        except ImportError:
            raise DPGFFormatError(
                "La librairie openpyxl n'est pas installee. "
                "Import Excel non disponible."
            )
        except Exception as e:
            raise DPGFFormatError(f"Erreur de lecture du fichier Excel: {str(e)}")

    def _parse_csv(
        self, content: bytes, mapping: DPGFColumnMappingDTO
    ) -> List[List[str]]:
        """Parse un fichier CSV.

        Detecte automatiquement le delimiteur (virgule, point-virgule, tab).

        Returns:
            Liste de lignes, chaque ligne etant une liste de valeurs string.
        """
        try:
            # Detecter l'encoding
            text = None
            for encoding in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if text is None:
                raise DPGFFormatError("Encoding du fichier CSV non reconnu")

            # Detecter le delimiteur
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
            reader = csv.reader(io.StringIO(text), dialect)

            rows = []
            for row_idx, row in enumerate(reader):
                if row_idx < mapping.ligne_debut:
                    continue
                if any(cell.strip() for cell in row):
                    rows.append(row)

            return rows

        except csv.Error as e:
            raise DPGFFormatError(f"Erreur de lecture du fichier CSV: {str(e)}")

    def _parse_ligne(
        self, row: List[str], mapping: DPGFColumnMappingDTO, ligne_num: int
    ) -> Optional[LigneDPGFDTO]:
        """Parse une ligne de donnees DPGF.

        Args:
            row: La ligne de donnees (liste de valeurs string).
            mapping: Le mapping des colonnes.
            ligne_num: Numero de ligne pour messages d'erreur.

        Returns:
            Un LigneDPGFDTO ou None si la ligne est vide/invalide.
        """
        def _get_col(idx: int) -> str:
            if idx < len(row):
                return row[idx].strip()
            return ""

        lot = _get_col(mapping.col_lot)
        description = _get_col(mapping.col_description)
        unite = _get_col(mapping.col_unite) or "U"
        quantite_raw = _get_col(mapping.col_quantite)
        pu_raw = _get_col(mapping.col_prix_unitaire)

        # Ignorer les lignes sans description
        if not description:
            return None

        # Lot par defaut si absent
        if not lot:
            lot = "DIVERS"

        # Parser les valeurs numeriques
        quantite, err_qty = _parse_decimal(quantite_raw, "quantite", ligne_num)
        prix_unitaire, err_pu = _parse_decimal(pu_raw, "prix_unitaire", ligne_num)

        # Les avertissements de parsing ne sont pas bloquants
        # (les valeurs sont mises a 0)

        return LigneDPGFDTO(
            lot=lot,
            description=description,
            unite=unite,
            quantite=quantite,
            prix_unitaire=prix_unitaire,
        )
